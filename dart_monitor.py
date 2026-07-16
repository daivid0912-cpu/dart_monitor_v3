# -*- coding: utf-8 -*-
"""
DART 메자닌/유상증자 신규발행 모니터링 봇
- 전환사채권발행결정(CB), 신주인수권부사채권발행결정(BW), 교환사채권발행결정(EB), 유상증자결정
- 유가증권/코스닥/코넥스만 (기타법인 제외)
- 기재정정/첨부정정 건 제외
- 자금조달목적 항목 금액 자동 합산
- 신규 공시 발견 시 이메일 알림 + 누적 엑셀 저장
"""

import os
import re
import json
import time
import smtplib
import traceback
from datetime import datetime
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

import pandas as pd
import OpenDartReader
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook

load_dotenv()

DART_API_KEY = os.getenv("DART_API_KEY")
EMAIL_FROM = os.getenv("EMAIL_FROM")
EMAIL_APP_PASSWORD = os.getenv("EMAIL_APP_PASSWORD")
EMAIL_TO = os.getenv("EMAIL_TO")
POLL_INTERVAL_SEC = int(os.getenv("POLL_INTERVAL_SEC", "600"))  # 기본 10분

SEEN_FILE = os.path.join(os.path.dirname(__file__), "seen_rcept.json")
OUTPUT_XLSX = os.path.join(os.path.dirname(__file__), "메자닌_발행사항_자동수집.xlsx")

if not DART_API_KEY:
    raise SystemExit("DART_API_KEY가 .env에 없습니다. .env.example을 참고해 .env를 만들어주세요.")

dart = OpenDartReader(DART_API_KEY)

CORP_CLS_MAP = {"Y": "코스피", "K": "코스닥", "N": "코넥스"}
ALLOWED_CORP_CLS = set(CORP_CLS_MAP.keys())  # E(기타법인) 제외

# 공시 제목 키워드 -> (내부 투자유형 라벨, dart.event()에 넘길 key_word)
TITLE_TO_TYPE = {
    "전환사채권발행결정": ("CB", "전환사채발행"),
    "신주인수권부사채권발행결정": ("BW", "신주인수권부사채발행"),
    "교환사채권발행결정": ("EB", "교환사채발행"),
    "유상증자결정": ("유상증자", "유상증자"),
}

EXCLUDE_PREFIXES = ["[기재정정]", "[첨부정정]", "기재정정", "첨부정정"]


def is_target_title(title: str):
    """제목이 모니터링 대상인지 확인. (정정 공시는 제외)"""
    for prefix in EXCLUDE_PREFIXES:
        if title.startswith(prefix):
            return None
    for keyword, meta in TITLE_TO_TYPE.items():
        if keyword in title:
            return meta
    return None


def load_seen():
    if os.path.exists(SEEN_FILE):
        with open(SEEN_FILE, "r", encoding="utf-8") as f:
            return set(json.load(f))
    return set()


def save_seen(seen: set):
    with open(SEEN_FILE, "w", encoding="utf-8") as f:
        json.dump(sorted(seen), f, ensure_ascii=False, indent=2)


def parse_document_extras(rcept_no: str):
    """공시 원문(document.xml)에서 Call option(%)과 대표주관회사를 best-effort로 추출."""
    try:
        raw = dart.document(rcept_no)
    except Exception as e:
        return {"call_option": None, "lead_manager": None, "parse_error": str(e)}

    text = re.sub(r"<[^>]+>", " ", raw)
    text = re.sub(r"&nbsp;?", " ", text)
    text = re.sub(r"\s+", " ", text)

    # 검증된 패턴: "매수인은 본 사채 발행가액의 NN%를 초과하여 매도청구권을 행사할 수 없다"
    # (실제 공시 원문 대조로 확인된 표현. "조기상환청구권"의 수익률과 반드시 구분해야 함 - 다른 항목임)
    call_option = None
    m = re.search(r"발행가액의\s*(\d{1,3}(?:\.\d{1,2})?)\s*%\s*를?\s*초과.{0,20}매도청구권", text)
    if not m:
        m = re.search(r"매도청구권.{0,20}발행가액의\s*(\d{1,3}(?:\.\d{1,2})?)\s*%", text)
    if m:
        call_option = m.group(1) + "%"

    lead_manager = None
    idx = text.find("대표주관회사")
    if idx != -1:
        window = text[idx: idx + 150]
        m = INSTITUTION_PATTERN.search(window)
        if m and len(m.group(0)) >= 3:
            lead_manager = m.group(0)

    participants_from_doc = extract_participant_table(text)

    return {
        "call_option": call_option,
        "lead_manager": lead_manager,
        "participants_from_doc": participants_from_doc,
    }


def extract_participant_table(text: str):
    """【특정인에 대한 대상자별 사채발행내역】 표(신탁업자+금액) +
    약칭 매핑 표(펀드번호->집합투자업자)를 함께 파싱해서,
    실질 투자자(집합투자업자) 기준으로 금액을 합산한 리스트를 반환.
    실제 공시 원문 구조 확인을 통해 검증된 로직."""
    from collections import defaultdict

    # 1) 펀드번호 -> 집합투자업자 매핑 (약칭 매핑 표)
    fund_to_manager = {}
    for m in re.finditer(
        r"본건\s*펀드(\d+)\s+.+?\s+([가-힣A-Za-z0-9\-]+(?:자산운용|투자운용)(?:투자일임)?(?:\(주\)|㈜)?)\s+"
        r"[가-힣A-Za-z0-9\-]+(?:증권|투자|은행)(?:\(주\)|㈜)?",
        text,
    ):
        fund_to_manager[m.group(1)] = m.group(2)

    # 2) 발행 대상자명 표: 기관명(+펀드번호 있으면) ... 선정 ... 금액
    amounts = defaultdict(int)
    for m in re.finditer(
        r"([가-힣A-Za-z0-9&\-\(\)㈜\s]+?)(?:\(본건\s*펀드(\d+)의\s*신탁업자\s*지위에서\))?"
        r"-\s*회사\s*경영상.*?선정-\s*([\d,]+)-",
        text,
    ):
        raw_name, fund_no, amount_str = m.group(1).strip(), m.group(2), m.group(3)
        try:
            amount_num = int(amount_str.replace(",", ""))
        except ValueError:
            continue
        if amount_num < 1_000_000:  # 너무 작은 값은 오탐 가능성 높아 제외
            continue
        display_name = fund_to_manager.get(fund_no, raw_name) if fund_no else raw_name
        display_name = display_name.replace("㈜", "").replace("(주)", "").strip()
        if len(display_name) < 2:
            continue
        amounts[display_name] += amount_num

    return [f"{name}({amt:,}원)" for name, amt in amounts.items()]


INSTITUTION_PATTERN = re.compile(
    r"[가-힣A-Za-z0-9&\-\(\)]*"
    r"(?:자산운용|증권|투자조합|신기술조합|신기술사업투자조합|벤처투자조합|파트너스|"
    r"캐피탈|자산투자|펀드|홀딩스|은행|화재|생명|저축은행|자산신탁|인베스트먼트|"
    r"투자일임|사모투자)"
)

AMOUNT_KEY_HINTS = ["amt", "금액"]
QTY_KEY_HINTS = ["qy", "cnt", "stk"]


def extract_participants_from_rows(rows: list):
    """상세 API의 각 row(취득자/인수인 1건에 대응한다고 가정)에서
    기관명 + 금액 또는 주식수를 best-effort로 추출."""
    participants = []
    for row in rows:
        row_dict = row.to_dict() if hasattr(row, "to_dict") else dict(row)

        inst_name = None
        for col, val in row_dict.items():
            val_str = str(val).strip()
            m = INSTITUTION_PATTERN.search(val_str)
            if m and len(m.group(0)) >= 3:
                inst_name = m.group(0)
                break

        if not inst_name:
            continue

        amount_val = None
        qty_val = None
        for col, val in row_dict.items():
            col_lower = col.lower()
            val_str = str(val).replace(",", "").strip()
            if not val_str or val_str in ("-", "nan"):
                continue
            try:
                num = int(val_str)
            except ValueError:
                continue
            if any(h in col_lower for h in AMOUNT_KEY_HINTS) and num > 0:
                amount_val = num
            elif any(h in col_lower for h in QTY_KEY_HINTS) and num > 0:
                qty_val = num

        if amount_val:
            participants.append(f"{inst_name}({amount_val:,}원)")
        elif qty_val:
            participants.append(f"{inst_name}({qty_val:,}주)")
        else:
            participants.append(inst_name)

    # 중복 제거 (같은 기관명+금액 조합이 반복될 수 있음)
    seen_p = set()
    deduped = []
    for p in participants:
        if p not in seen_p:
            seen_p.add(p)
            deduped.append(p)
    return deduped


def sum_fdpp_fields(row: pd.Series):
    """자금조달목적(fdpp_ 로 시작하는 필드) 항목 금액을 모두 합산.
    필드명이 다를 경우를 대비해 '금액'을 담고 있을 만한 숫자형 필드도 함께 로그로 남김."""
    total = 0
    matched_fields = {}
    for col, val in row.items():
        if col.startswith("fdpp_"):
            try:
                amount = int(str(val).replace(",", "").strip() or 0)
            except ValueError:
                amount = 0
            matched_fields[col] = amount
            total += amount
    return total, matched_fields


def classify_equity_type(row: pd.Series):
    """유상증자결정 건에서 보통주식/기타주식 여부 best-effort 판별.
    실제 필드명이 다를 수 있어 방어적으로 여러 후보 필드명을 확인."""
    candidates_common = ["nstk_ostk_cnt", "nstk_ostk_cnt2"]
    candidates_other = ["nstk_estk_cnt", "nstk_estk_cnt2"]

    def get_num(colnames):
        for c in colnames:
            if c in row and str(row[c]).strip() not in ("", "-"):
                try:
                    return int(str(row[c]).replace(",", "").strip())
                except ValueError:
                    continue
        return 0

    common_cnt = get_num(candidates_common)
    other_cnt = get_num(candidates_other)

    if other_cnt > 0 and common_cnt > 0:
        return "유상증자(보통+기타)"
    elif other_cnt > 0:
        return "유상증자(기타주식)"
    elif common_cnt > 0:
        return "유상증자(보통주식)"
    else:
        return "유상증자(확인필요)"


def fetch_new_filings(start_date: str = None, end_date: str = None):
    """지정한 기간(start_date~end_date)의 주요사항보고(kind='B') 공시 중
    대상 필터를 통과한 신규 건 반환. 날짜 미지정 시 전날(어제) 하루치."""
    from datetime import timedelta

    if not start_date and not end_date:
        target = (datetime.today() - timedelta(days=1)).strftime("%Y-%m-%d")
        start_date = end_date = target
    elif start_date and not end_date:
        end_date = start_date
    elif end_date and not start_date:
        start_date = end_date

    df = dart.list(start=start_date, end=end_date, kind="B", final=True)

    if df is None or len(df) == 0:
        return []

    results = []
    for _, r in df.iterrows():
        corp_cls = r.get("corp_cls")
        if corp_cls not in ALLOWED_CORP_CLS:
            continue

        title = str(r.get("report_nm", ""))
        meta = is_target_title(title)
        if not meta:
            continue

        results.append(
            {
                "rcept_no": r.get("rcept_no"),
                "rcept_dt": r.get("rcept_dt"),
                "corp_code": r.get("corp_code"),
                "corp_name": r.get("corp_name"),
                "corp_cls": corp_cls,
                "report_nm": title,
                "invest_type": meta[0],
                "event_keyword": meta[1],
            }
        )
    return results


def enrich_filing(filing: dict):
    """dart.event()로 상세정보 조회 후 자금조달목적 합산 + 유상증자 종류 판별."""
    try:
        detail_df = dart.event(
            filing["corp_code"],
            filing["event_keyword"],
            start=filing["rcept_dt"],
            end=filing["rcept_dt"],
        )
    except Exception as e:
        filing["fund_amount"] = None
        filing["fund_detail"] = {}
        filing["equity_type"] = None
        filing["participants"] = []
        filing["call_option"] = None
        filing["lead_manager"] = None
        filing["detail_error"] = str(e)
        return filing

    if detail_df is None or len(detail_df) == 0:
        filing["fund_amount"] = None
        filing["fund_detail"] = {}
        filing["equity_type"] = None
        filing["participants"] = []
        filing["call_option"] = None
        filing["lead_manager"] = None
        return filing

    # rcept_no로 매칭되는 행 전체 찾기 (참여기관이 여러 개면 여러 행일 수 있음)
    match_col = "rcept_no" if "rcept_no" in detail_df.columns else None
    if match_col:
        matched = detail_df[detail_df[match_col] == filing["rcept_no"]]
        matched_rows = [r for _, r in matched.iterrows()] if len(matched) > 0 else [detail_df.iloc[0]]
    else:
        matched_rows = [detail_df.iloc[0]]

    row = matched_rows[0]  # 자금조달목적/주식종류 등 필링 레벨 정보는 첫 행 기준

    total, matched_fields = sum_fdpp_fields(row)
    filing["fund_amount"] = total if total > 0 else None
    filing["fund_detail"] = matched_fields

    if filing["invest_type"] == "유상증자":
        filing["equity_type"] = classify_equity_type(row)
    else:
        filing["equity_type"] = None

    filing["participants"] = extract_participants_from_rows(matched_rows)
    filing["all_rows_raw"] = [r.to_dict() for r in matched_rows]  # 디버깅용: 전체 행

    extras = parse_document_extras(filing["rcept_no"])
    filing["call_option"] = extras.get("call_option")
    filing["lead_manager"] = extras.get("lead_manager")
    filing["extras_debug"] = extras

    # 참여기관: 구조화 API 기반(row)과 원문 텍스트 기반(doc) 결과를 합쳐서 중복 제거
    doc_participants = extras.get("participants_from_doc", [])
    if doc_participants:
        combined = list(filing.get("participants", [])) + doc_participants
        seen_c = set()
        deduped_c = []
        for p in combined:
            if p not in seen_c:
                seen_c.add(p)
                deduped_c.append(p)
        filing["participants"] = deduped_c

    filing["raw_row"] = row.to_dict()  # 디버깅/수동확인용 원본 저장 (첫 행)
    return filing


def append_to_excel(filings: list):
    if not filings:
        return

    if os.path.exists(OUTPUT_XLSX):
        wb = load_workbook(OUTPUT_XLSX)
        ws = wb.active
        next_no = ws.max_row  # 헤더 1행 있으므로 max_row가 다음 No와 일치
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "메자닌 발행사항"
        headers = [
            "No", "발행일자", "발행대상", "거래소", "투자유형",
            "발행규모(원)", "참여기관(참여금액)", "Call option(%)",
            "시가총액(억원)", "주관기관", "원문링크",
        ]
        ws.append(headers)
        next_no = 1

    for f in filings:
        invest_type_label = f["equity_type"] if f["equity_type"] else f["invest_type"]
        link = f"https://dart.fss.or.kr/dsaf001/main.do?rcpNo={f['rcept_no']}"
        participants_str = "\n".join(f.get("participants", [])) if f.get("participants") else ""
        ws.append(
            [
                next_no,
                f["rcept_dt"],
                f["corp_name"],
                CORP_CLS_MAP.get(f["corp_cls"], f["corp_cls"]),
                invest_type_label,
                f.get("fund_amount"),
                participants_str,
                f.get("call_option") or "",
                "",  # 시가총액 - 수동 확인
                f.get("lead_manager") or "",
                link,
            ]
        )
        next_no += 1

    wb.save(OUTPUT_XLSX)


def send_email(filings: list):
    if not filings:
        return
    if not (EMAIL_FROM and EMAIL_APP_PASSWORD and EMAIL_TO):
        print("[경고] 이메일 설정이 없어 알림을 건너뜁니다. (.env의 EMAIL_* 확인)")
        return

    lines = []
    for f in filings:
        invest_type_label = f["equity_type"] if f["equity_type"] else f["invest_type"]
        amount = f.get("fund_amount")
        amount_str = f"{amount:,}원" if amount else "금액 확인 필요"
        link = f"https://dart.fss.or.kr/dsaf001/main.do?rcpNo={f['rcept_no']}"
        lines.append(
            f"- [{invest_type_label}] {f['corp_name']} ({CORP_CLS_MAP.get(f['corp_cls'])})"
            f"\n  발행규모(자금조달목적 합산): {amount_str}"
            f"\n  공시일: {f['rcept_dt']}"
            f"\n  원문: {link}\n"
        )

    body = f"신규 메자닌/유상증자 공시 {len(filings)}건 발견\n\n" + "\n".join(lines)

    msg = MIMEMultipart()
    msg["From"] = EMAIL_FROM
    msg["To"] = EMAIL_TO
    msg["Subject"] = f"[DART 알림] 신규 메자닌/유상증자 공시 {len(filings)}건"
    msg.attach(MIMEText(body, "plain", "utf-8"))

    with smtplib.SMTP("smtp.gmail.com", 587) as server:
        server.starttls()
        server.login(EMAIL_FROM, EMAIL_APP_PASSWORD)
        server.sendmail(EMAIL_FROM, EMAIL_TO, msg.as_string())

    print(f"[이메일 발송 완료] {len(filings)}건")


def run_once(seen: set, debug=False, start_date=None, end_date=None, ignore_seen=False):
    candidates = fetch_new_filings(start_date=start_date, end_date=end_date)
    if ignore_seen:
        new_filings = candidates
    else:
        new_filings = [c for c in candidates if c["rcept_no"] not in seen]

    if not new_filings:
        print(f"[{datetime.now().strftime('%H:%M:%S')}] 신규 공시 없음 (전날 대상 공시 {len(candidates)}건 확인)")
        return seen

    enriched = []
    for f in new_filings:
        enriched.append(enrich_filing(f))
        time.sleep(0.3)  # API 과호출 방지

    if debug:
        for f in enriched:
            print("=" * 60)
            print(f"{f['corp_name']} / {f['invest_type']} / rcept_no={f['rcept_no']}")
            print("자금조달목적 필드:", f["fund_detail"])
            print("합산 발행규모:", f.get("fund_amount"))
            if f["invest_type"] == "유상증자":
                print("주식종류 판별:", f["equity_type"])
            print("Call option 파싱 결과:", f.get("call_option"))
            print("대표주관회사 파싱 결과:", f.get("lead_manager"))
            extras_dbg = f.get("extras_debug", {})
            print("원문에서 뽑은 참여기관(합산 후):", extras_dbg.get("participants_from_doc"))
            print("추출된 참여기관:", f.get("participants"))
            print("전체 행 원본 (참여기관 필드명 검증용):")
            print(json.dumps(f.get("all_rows_raw", []), ensure_ascii=False, indent=2))
            print("원본 필드 전체(첫 조회 시 필드명 검증용):")
            print(json.dumps(f.get("raw_row", {}), ensure_ascii=False, indent=2))

    append_to_excel(enriched)
    send_email(enriched)

    for f in new_filings:
        seen.add(f["rcept_no"])
    save_seen(seen)

    print(f"[{datetime.now().strftime('%H:%M:%S')}] 신규 공시 {len(new_filings)}건 처리 완료")
    return seen


def main():
    import sys
    import argparse

    parser = argparse.ArgumentParser()
    parser.add_argument("--once", action="store_true")
    parser.add_argument("--debug", action="store_true")
    parser.add_argument("--start", type=str, default=None, help="조회 시작일 YYYY-MM-DD")
    parser.add_argument("--end", type=str, default=None, help="조회 종료일 YYYY-MM-DD")
    parser.add_argument("--ignore-seen", action="store_true", help="이미 처리한 건도 다시 포함")
    args, _ = parser.parse_known_args()

    debug = args.debug
    once = args.once
    # GitHub Actions workflow_dispatch 입력값은 환경변수로 전달됨
    start_date = args.start or os.getenv("TARGET_START_DATE") or None
    end_date = args.end or os.getenv("TARGET_END_DATE") or None
    ignore_seen = args.ignore_seen or (os.getenv("IGNORE_SEEN", "").lower() == "true")

    seen = load_seen()
    print(f"DART 모니터링 시작 (debug={debug}, start={start_date}, end={end_date}, ignore_seen={ignore_seen})")

    if once or start_date or end_date:
        run_once(seen, debug=debug, start_date=start_date, end_date=end_date, ignore_seen=ignore_seen)
        return

    while True:
        try:
            seen = run_once(seen, debug=debug)
        except Exception:
            print("[에러 발생]")
            traceback.print_exc()
        time.sleep(POLL_INTERVAL_SEC)


if __name__ == "__main__":
    main()
